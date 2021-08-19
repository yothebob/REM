from openpyxl import Workbook
from openpyxl import load_workbook
import sqlite3

workbook = load_workbook('part_pricing_list.xlsx',data_only=True)

parts_list = workbook["part_pricing_list"]

name= []
des = []
price = []

db_name = "parts_list.db"

con = sqlite3.connect(db_name)
cur = con.cursor()

for row in parts_list.iter_rows(min_row=1,max_row=995,min_col=3,max_col=3,values_only=True):
    res = ''.join(map(str,row))
    name.append(res)

for row in parts_list.iter_rows(min_row=1,max_row=995,min_col=4,max_col=4,values_only=True):
    res = ''.join(map(str,row))
    des.append(res)


for row in parts_list.iter_rows(min_row=1,max_row=995,min_col=7,max_col=7,values_only=True):
    res = ''.join(map(str,row))
    try:
        price.append(float(res))
    except:
        price.append(0)

print(price)


cur.execute('''CREATE TABLE IF NOT EXISTS inventory
              (id real PRIMARY KEY, name text, des text, price real)''')

for num in range(len(name)):
    con.execute("insert into inventory values (?, ?, ?, ?)", (num,name[num],des[num],int(price[num])))

con.commit()

con.close()

